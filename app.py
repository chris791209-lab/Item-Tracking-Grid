import streamlit as st
import pandas as pd
import io
import os
import tempfile
import zipfile
import openpyxl
from PIL import Image

# ==========================================
# 0. 頁面基本設定 & 密碼管控
# ==========================================
st.set_page_config(page_title="D240 Item Tracking Grid Generator", layout="wide")

def check_password():
    def password_entered():
        if st.session_state["password"] == st.secrets["app_password"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        st.title("🔒 系統登入")
        st.text_input("🔒 請輸入 AE 部門共用密碼以啟用工具：", type="password", on_change=password_entered, key="password")
        return False
    elif not st.session_state["password_correct"]:
        st.title("🔒 系統登入")
        st.text_input("🔒 請輸入 AE 部門共用密碼以啟用工具：", type="password", on_change=password_entered, key="password")
        st.error("❌ 密碼錯誤，請重新輸入。")
        return False
    else:
        return True

if not check_password():
    st.stop()

# ==========================================
# 1. 主程式標題 
# ==========================================
st.title("🎃 D240 Item Tracking Grid自動生成工具")
st.markdown("請上傳您的專案檔案 (Program Sheet 或 Data 表)。系統會智慧解析內容，並透過 DPCI 完美匹配您上傳的圖片 ZIP 包！")

# ==========================================
# 2. 檔案上傳區塊
# ==========================================
st.markdown("### 📄 步驟 1：上傳資料檔案")
uploaded_files = st.file_uploader("📁 請將 [Program Sheet] 與 [Data 表] 一同拖曳至此", 
                                  type=["xlsx", "xls", "csv"], 
                                  accept_multiple_files=True)

st.markdown("### 🖼️ 步驟 2：上傳圖片 ZIP 壓縮檔 (一勞永逸配對法)")
st.info("💡 將所有圖片放入 ZIP 壓縮檔。**圖片檔名請設定為 DPCI**（例：`240-02-1234.png` 或 `240021234.jpg`）。")
uploaded_zip = st.file_uploader("📁 請上傳 .zip 圖片壓縮檔", type=["zip"])

st.divider() 

# 🛠️ 工具函式
def clean_string(val):
    return str(val).replace(' ', '').replace(':', '').replace('#', '').replace('\n', '').replace('\r', '').upper()

def clean_dpci_for_map(d):
    return str(d).replace('-', '').strip()

# 💡 圖片白底轉換函式 (解決去背圖變黑的問題)
def add_white_background(img):
    if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
        img = img.convert('RGBA')
        bg = Image.new('RGB', img.size, (255, 255, 255))
        bg.paste(img, mask=img.split()[3]) # 利用 Alpha 通道當遮罩，把商品貼到白底上
        return bg
    elif img.mode != 'RGB':
        return img.convert('RGB')
    return img

# ==========================================
# 3. 核心處理邏輯
# ==========================================
if uploaded_files:
    if st.button("✨ 智慧生成 Item Tracking Grid", type="primary"):
        with st.spinner("雙引擎聯合解析與圖片配對中，請稍候..."):
            with tempfile.TemporaryDirectory() as temp_dir:
                try:
                    # --- 1. 解壓縮圖片包 ---
                    if uploaded_zip:
                        zip_io = io.BytesIO(uploaded_zip.getvalue())
                        with zipfile.ZipFile(zip_io, 'r') as zip_ref:
                            zip_ref.extractall(temp_dir)

                    tabular_items = []  
                    parsed_items = []   
                    
                    cat_mapping = {}
                    fact_mapping = {}
                    qty_mapping = {}
                    desc_mapping = {}

                    # --- 2. 聯合讀取所有檔案 ---
                    for file in uploaded_files:
                        file_ext = file.name.lower()
                        file_bytes = file.getvalue()
                        
                        # (A) 處理 CSV 格式
                        if file_ext.endswith('.csv'):
                            try:
                                df = pd.read_csv(io.BytesIO(file_bytes), header=None, encoding='utf-8')
                            except UnicodeDecodeError:
                                df = pd.read_csv(io.BytesIO(file_bytes), header=None, encoding='cp1252', errors='replace')
                            
                            header_idx = -1
                            for i in range(min(20, len(df))):
                                row_vals = [str(v).strip().upper() for v in df.iloc[i].values]
                                if any('DPCI' in v for v in row_vals) and (any('DESCRIPTION' in v for v in row_vals) or any('VENDOR' in v for v in row_vals) or any('SUBCLASS' in v for v in row_vals)):
                                    header_idx = i; break
                                    
                            if header_idx != -1:
                                df.columns = [str(c).strip().upper() for c in df.iloc[header_idx]]
                                df = df.loc[:, ~df.columns.duplicated(keep='first')]
                                df = df.iloc[header_idx + 1:].reset_index(drop=True)
                                
                                def get_col(candidates):
                                    for c in candidates:
                                        for col in df.columns:
                                            if c in col.replace(' ', '').replace('_', ''): return col
                                    return None
                                    
                                dpci_col = get_col(['DPCI'])
                                desc_col = get_col(['PRODUCTDESCRIPTION', 'ITEMDESC', 'DESCRIPTION'])
                                fact_col = get_col(['PRODUCTBUSINESSPARTNER', 'IMPORTVENDORNAME', 'VENDOR', 'FACTORY'])
                                qty_col = get_col(['TOTALUNITS', 'ENTTTLRCPTU', 'QTY', 'CASEPACK'])
                                cat_col = get_col(['SUBCLASSNAME', 'CATEGORY'])
                                
                                if dpci_col:
                                    for _, row in df.iterrows():
                                        raw_dpci = str(row[dpci_col]).strip()
                                        if raw_dpci.lower() in ['nan', 'none', '']: continue
                                        dpci_key = clean_dpci_for_map(raw_dpci)
                                        if dpci_key.endswith('.0'): dpci_key = dpci_key[:-2]
                                        
                                        desc = str(row[desc_col]).strip() if desc_col else ""
                                        fact = str(row[fact_col]).strip() if fact_col else ""
                                        qty = str(row[qty_col]).strip() if qty_col else ""
                                        cat = str(row[cat_col]).strip() if cat_col else ""
                                        
                                        if desc and desc.lower() != 'nan': desc_mapping[dpci_key] = desc
                                        if fact and fact.lower() != 'nan': fact_mapping[dpci_key] = fact
                                        if qty and qty.lower() != 'nan': qty_mapping[dpci_key] = qty
                                        if cat and cat.lower() != 'nan': cat_mapping[dpci_key] = cat
                                        
                                        tabular_items.append({'DPCI': raw_dpci, 'ITEM_DESC': desc if desc.lower() != 'nan' else "", 'Factory Name': fact if fact.lower() != 'nan' else "", 'Factory ID': "", 'QTY': qty if qty.lower() != 'nan' else ""})

                        # (B) 處理 Excel 格式
                        elif file_ext.endswith(('.xlsx', '.xls')):
                            xls = pd.ExcelFile(io.BytesIO(file_bytes))
                            for sheet_name in xls.sheet_names:
                                df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                                
                                header_idx = -1
                                for i in range(min(20, len(df))):
                                    row_vals = [str(v).strip().upper() for v in df.iloc[i].values]
                                    if any('DPCI' in v for v in row_vals) and (any('DESCRIPTION' in v for v in row_vals) or any('VENDOR' in v for v in row_vals) or any('SUBCLASS' in v for v in row_vals)):
                                        header_idx = i; break
                                
                                is_tabular = False
                                if header_idx != -1:
                                    is_tabular = True
                                    df.columns = [str(c).strip().upper() for c in df.iloc[header_idx]]
                                    df = df.loc[:, ~df.columns.duplicated(keep='first')]
                                    df = df.iloc[header_idx + 1:].reset_index(drop=True)
                                    
                                    def get_col(candidates):
                                        for c in candidates:
                                            for col in df.columns:
                                                if c in col.replace(' ', '').replace('_', ''): return col
                                        return None
                                        
                                    dpci_col = get_col(['DPCI'])
                                    desc_col = get_col(['PRODUCTDESCRIPTION', 'ITEMDESC', 'DESCRIPTION'])
                                    fact_col = get_col(['PRODUCTBUSINESSPARTNER', 'IMPORTVENDORNAME', 'VENDOR', 'FACTORY'])
                                    qty_col = get_col(['TOTALUNITS', 'ENTTTLRCPTU', 'QTY', 'CASEPACK'])
                                    cat_col = get_col(['SUBCLASSNAME', 'CATEGORY'])
                                    
                                    if dpci_col:
                                        for _, row in df.iterrows():
                                            raw_dpci = str(row[dpci_col]).strip()
                                            if raw_dpci.lower() in ['nan', 'none', '']: continue
                                            dpci_key = clean_dpci_for_map(raw_dpci)
                                            if dpci_key.endswith('.0'): dpci_key = dpci_key[:-2]
                                            
                                            desc = str(row[desc_col]).strip() if desc_col else ""
                                            fact = str(row[fact_col]).strip() if fact_col else ""
                                            qty = str(row[qty_col]).strip() if qty_col else ""
                                            cat = str(row[cat_col]).strip() if cat_col else ""
                                            
                                            if desc and desc.lower() != 'nan': desc_mapping[dpci_key] = desc
                                            if fact and fact.lower() != 'nan': fact_mapping[dpci_key] = fact
                                            if qty and qty.lower() != 'nan': qty_mapping[dpci_key] = qty
                                            if cat and cat.lower() != 'nan': cat_mapping[dpci_key] = cat
                                            
                                            tabular_items.append({'DPCI': raw_dpci, 'ITEM_DESC': desc if desc.lower() != 'nan' else "", 'Factory Name': fact if fact.lower() != 'nan' else "", 'Factory ID': "", 'QTY': qty if qty.lower() != 'nan' else ""})

                                # 若非資料表，啟動卡片排版掃描器
                                if not is_tabular:
                                    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                                    sheet = wb[sheet_name]
                                    for r in range(1, sheet.max_row + 1):
                                        for c in range(1, sheet.max_column + 1):
                                            cell_val = str(sheet.cell(row=r, column=c).value or "").strip()
                                            if not cell_val: continue
                                            
                                            val_clean = clean_string(cell_val)
                                            dpci = ""
                                            
                                            if val_clean == 'DPCI':
                                                for offset in range(1, 5):
                                                    if c+offset > sheet.max_column: break
                                                    v = str(sheet.cell(row=r, column=c+offset).value or "").strip()
                                                    if v and v.lower() != 'none':
                                                        dpci = v; break
                                            elif val_clean.startswith('DPCI') and len(val_clean) > 4:
                                                dpci = cell_val.split(':')[-1].replace('#', '').strip()
                                                
                                            if dpci:
                                                desc = ""
                                                qty = ""
                                                factory_name = ""
                                                factory_id = ""
                                                
                                                for i in range(1, 20):
                                                    if r+i > sheet.max_row: break
                                                    if not desc:
                                                        c_val = clean_string(sheet.cell(row=r+i, column=c).value)
                                                        if 'DESCRIPTION' in c_val:
                                                            for offset in range(1, 5):
                                                                if c+offset > sheet.max_column: break
                                                                v = str(sheet.cell(row=r+i, column=c+offset).value or "").strip()
                                                                if v and v.lower() != 'none': desc = v; break
                                                    if not qty:
                                                        for j in range(c, min(c+12, sheet.max_column + 1)):
                                                            q_val = clean_string(sheet.cell(row=r+i, column=j).value)
                                                            if 'QTY' in q_val or 'CASEPACK' in q_val:
                                                                for offset in range(1, 5):
                                                                    if j+offset > sheet.max_column: break
                                                                    v = str(sheet.cell(row=r+i, column=j+offset).value or "").strip()
                                                                    if v and v.lower() != 'none':
                                                                        qty = v
                                                                        if qty.endswith('.0'): qty = qty[:-2]
                                                                        break
                                                                break
                                                    if not factory_name:
                                                        f_val = clean_string(sheet.cell(row=r+i, column=c).value)
                                                        if 'FACTORY' in f_val or 'VENDOR' in f_val:
                                                            raw_fact = str(sheet.cell(row=r+i, column=c).value or "").strip().replace('"', '')
                                                            if len(raw_fact) < 10 or raw_fact.endswith(':'):
                                                                for offset in range(1, 5):
                                                                    if c+offset > sheet.max_column: break
                                                                    v = str(sheet.cell(row=r+i, column=c+offset).value or "").strip()
                                                                    if v and v.lower() != 'none': raw_fact = v; break
                                                            if ':' in raw_fact: raw_fact = raw_fact.split(':', 1)[-1].strip()
                                                            parts = raw_fact.split('/')
                                                            if len(parts) >= 1: factory_name = parts[0].strip()
                                                            if len(parts) >= 2: factory_id = parts[1].strip()

                                                parsed_items.append({
                                                    'DPCI': dpci,
                                                    'ITEM_DESC': desc,
                                                    'Factory Name': factory_name,
                                                    'Factory ID': factory_id,
                                                    'QTY': qty
                                                })

                    # --- 3. 決策資料來源 ---
                    if len(parsed_items) > 0:
                        final_items = parsed_items
                    else:
                        unique_tabular = []
                        seen_dpcis = set()
                        for item in tabular_items:
                            c_dpci = clean_dpci_for_map(item['DPCI'])
                            if c_dpci not in seen_dpcis:
                                seen_dpcis.add(c_dpci)
                                unique_tabular.append(item)
                        final_items = unique_tabular

                    if not final_items:
                        st.warning("⚠️ 無法在檔案中找到任何有效資料。請確認檔案格式！")
                        st.stop()
                    
                    df_out = pd.DataFrame(final_items)

                    # --- 4. 終極 VLOOKUP 補齊資料 ---
                    df_out['CATEGORY'] = df_out['DPCI'].apply(lambda x: cat_mapping.get(clean_dpci_for_map(x), ""))
                    df_out['Factory Name'] = df_out.apply(lambda r: fact_mapping.get(clean_dpci_for_map(r['DPCI']), "") if not r['Factory Name'] else r['Factory Name'], axis=1)
                    df_out['QTY'] = df_out.apply(lambda r: qty_mapping.get(clean_dpci_for_map(r['DPCI']), "") if not r['QTY'] else r['QTY'], axis=1)
                    df_out['ITEM_DESC'] = df_out.apply(lambda r: desc_mapping.get(clean_dpci_for_map(r['DPCI']), "") if not r['ITEM_DESC'] else r['ITEM_DESC'], axis=1)

                    df_out['Factory Name'] = df_out['Factory Name'].replace('', 'Unknown Factory')
                    df_out['Factory Name'] = df_out['Factory Name'].fillna('Unknown Factory')

                    df_out.sort_values(by='Factory Name', inplace=True)
                    df_out.reset_index(drop=True, inplace=True)

                    target_columns = [
                        "DPCI", "CATEGORY", "ITEM_DESC", "PHOTO", "FRP Level", 
                        "Red Seal(Y/N)", "CF item( Y/N )", "Tollgate Exempt", 
                        "TPR Lite/Exempt", "Factory Name", "Factory ID", 
                        "Total SKU per Factory", "QTY", 
                        "Tollgate Date", "TPR Date", "Dupro Date", "Result", 
                        "TOP Result", "FRI plan", "Port of Export", 
                        "1st Ship window", "Inspection Office"
                    ]
                    for col in target_columns:
                        if col not in df_out.columns: df_out[col] = ""

                    # --- 5. XlsxWriter 排版與 ZIP 圖片綁定 ---
                    output = io.BytesIO()
                    img_insert_count = 0
                    
                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        workbook = writer.book
                        worksheet = workbook.add_worksheet('Program Items')
                        
                        format_cache = {}
                        def get_fmt(t=1, b=1, l=1, r=1, align=None):
                            key = (t, b, l, r, align)
                            if key not in format_cache:
                                props = {'font_name': 'Arial', 'valign': 'vcenter', 'text_wrap': True, 'top': t, 'bottom': b, 'left': l, 'right': r}
                                if align: props['align'] = align
                                format_cache[key] = workbook.add_format(props)
                            return format_cache[key]

                        fmt_total = workbook.add_format({'font_name': 'Arial', 'bold': True, 'font_color': '#800080', 'font_size': 12, 'align': 'center', 'valign': 'vcenter'})
                        worksheet.write(0, 11, len(df_out), fmt_total)
                        
                        fmt_header = workbook.add_format({'bold': True, 'bg_color': '#FFD966', 'border': 1, 'font_name': 'Arial', 'valign': 'vcenter', 'text_wrap': True}) 
                        for c, col_name in enumerate(target_columns):
                            worksheet.write(1, c, col_name, fmt_header)
                            worksheet.set_column(c, c, 15) 
                            
                        worksheet.set_column(3, 3, 16) 
                        worksheet.set_column(9, 9, 25) 
                        worksheet.set_column(10, 10, 15) 
                        
                        factories = df_out['Factory Name'].tolist()
                        groups = []
                        if factories:
                            start_idx = 0
                            for i in range(1, len(factories)):
                                if factories[i] != factories[start_idx]:
                                    groups.append((start_idx, i - 1, factories[start_idx]))
                                    start_idx = i
                            groups.append((start_idx, len(factories) - 1, factories[start_idx]))

                        excel_row_offset = 2 
                        for s_idx, e_idx, f_name in groups:
                            f_id = str(df_out.iloc[s_idx]['Factory ID']).strip()
                            
                            for i in range(s_idx, e_idx + 1):
                                excel_row = i + excel_row_offset
                                worksheet.set_row(excel_row, 80) 
                                
                                dpci_val = str(df_out.iloc[i]['DPCI']).strip()
                                safe_name = "".join(x for x in dpci_val if x.isalnum() or x in "-_")
                                if safe_name.endswith('.0'): safe_name = safe_name[:-2]
                                clean_dpci_img = safe_name.replace('-', '')
                                
                                # 🔍 在 ZIP 中精準比對檔名
                                img_path = None
                                search_names = [
                                    f"{safe_name}.png", f"{safe_name}.jpg", f"{safe_name}.jpeg",
                                    f"{dpci_val.lower()}.png", f"{dpci_val.lower()}.jpg", f"{dpci_val.lower()}.jpeg",
                                    f"{clean_dpci_img}.png", f"{clean_dpci_img}.jpg", f"{clean_dpci_img}.jpeg"
                                ]
                                for root, dirs, files in os.walk(temp_dir):
                                    for file in files:
                                        if file.lower() in [s.lower() for s in search_names]:
                                            img_path = os.path.join(root, file)
                                            break
                                    if img_path: break
                                
                                for c, col_name in enumerate(target_columns):
                                    if c in [9, 10]: continue 
                                    
                                    t_border = 2 if i == s_idx else 1
                                    b_border = 2 if i == e_idx else 1
                                    l_border = 2 if c == 0 else 1
                                    r_border = 2 if c == len(target_columns) - 1 else 1
                                    
                                    fmt = get_fmt(t_border, b_border, l_border, r_border)
                                    
                                    if c == 3: 
                                        worksheet.write(excel_row, c, "", fmt)
                                        if img_path:
                                            try:
                                                with Image.open(img_path) as img:
                                                    # 👉 呼叫白底轉換函式，消除去背圖變黑的問題
                                                    img = add_white_background(img)
                                                    img.thumbnail((100, 100))
                                                    resized_path = os.path.join(temp_dir, f"resized_{i}.png")
                                                    img.save(resized_path, "PNG")
                                                    worksheet.insert_image(excel_row, c, resized_path, {'x_offset': 5, 'y_offset': 5})
                                                    img_insert_count += 1
                                            except: pass
                                            
                                    elif c == 11: 
                                        if i == s_idx: worksheet.write(excel_row, c, e_idx - s_idx + 1, fmt)
                                        else: worksheet.write(excel_row, c, "", fmt)
                                    else:
                                        worksheet.write(excel_row, c, df_out.iloc[i][col_name], fmt)
                                        
                            fmt_fact = get_fmt(t=2, b=2, l=1, r=1, align='center')
                            if s_idx == e_idx:
                                worksheet.write(s_idx + excel_row_offset, 9, f_name, fmt_fact)
                                worksheet.write(s_idx + excel_row_offset, 10, f_id, fmt_fact)
                            else:
                                worksheet.merge_range(s_idx + excel_row_offset, 9, e_idx + excel_row_offset, 9, f_name, fmt_fact)
                                worksheet.merge_range(s_idx + excel_row_offset, 10, e_idx + excel_row_offset, 10, f_id, fmt_fact)

                    processed_data = output.getvalue()
                    
                    st.success(f"✅ 處理完成！共解析出 **{len(df_out)}** 筆商品，並從 ZIP 檔中成功置入 **{img_insert_count}** 張完美白底圖片。")
                    
                    st.download_button(
                        label="📥 下載 Item Tracking Grid.xlsx",
                        data=processed_data,
                        file_name="Item_Tracking_Grid.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                except Exception as e:
                    st.error(f"❌ 處理檔案時發生錯誤: {str(e)}")

else:
    st.info("💡 提示：請在上方直接拖曳上傳您的專案 Excel/CSV 檔案與圖片 ZIP 檔。")
